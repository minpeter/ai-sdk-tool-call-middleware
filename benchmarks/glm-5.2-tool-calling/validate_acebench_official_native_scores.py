#!/usr/bin/env python3
"""Strictly validate direct-FC ACEBench scoring and official summaries."""

from __future__ import annotations

import argparse
import hashlib
import json
import subprocess
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


NORMAL_CATEGORIES = {
    "normal_single_turn_single_function",
    "normal_single_turn_parallel_function",
    "normal_multi_turn_user_adjust",
    "normal_multi_turn_user_switch",
    "normal_similar_api",
    "normal_preference",
    "normal_atom_bool",
    "normal_atom_enum",
    "normal_atom_number",
    "normal_atom_list",
    "normal_atom_object_deep",
    "normal_atom_object_short",
}
SPECIAL_CATEGORIES = {
    "special_incomplete",
    "special_error_param",
    "special_irrelevant",
}
AGENT_CATEGORIES = {"agent_multi_step", "agent_multi_turn"}
ALL_CATEGORIES = NORMAL_CATEGORIES | SPECIAL_CATEGORIES | AGENT_CATEGORIES
ARMS = ("glm52-native-FC", "glm52-native-plus-FC")
LANGUAGES = ("en", "zh")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise RuntimeError(f"{path}: expected object")
    return value


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise RuntimeError(f"{path}:{line_number}: expected object")
            rows.append(value)
    return rows


def category_from_path(path: Path) -> str:
    prefix, suffix = "data_", "_result.json"
    if not path.name.startswith(prefix) or not path.name.endswith(suffix):
        raise RuntimeError(f"unexpected result filename: {path}")
    return path.name[len(prefix) : -len(suffix)]


def workbook_summaries(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    model_index = headers.index("Model")
    summary_index = headers.index("Summary")
    output: dict[str, float] = {}
    for row in rows[1:]:
        model, score = row[model_index], row[summary_index]
        if not isinstance(model, str) or not isinstance(score, (int, float)):
            raise RuntimeError(f"{path}: invalid workbook row")
        output[model] = float(score)
    return output


def canonical_tree_hash(root: Path, paths: list[Path]) -> str:
    records = [
        {"path": str(path.relative_to(root)), "sha256": sha256_file(path)}
        for path in sorted(paths)
    ]
    encoded = json.dumps(
        records, separators=(",", ":"), sort_keys=True
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ace-root", required=True, type=Path)
    parser.add_argument("--scoring-root", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()

    ace_root = args.ace_root.resolve()
    scoring_root = args.scoring_root.resolve()
    output = args.out.resolve()
    if output.exists():
        raise RuntimeError(f"refusing to overwrite validation: {output}")
    manifest = load_json(scoring_root / "scoring-manifest.json")
    revision = subprocess.check_output(
        ["git", "-C", str(ace_root), "rev-parse", "HEAD"], text=True
    ).strip()
    if revision != manifest.get("officialCommit"):
        raise RuntimeError("ACEBench official commit drift")
    scorer = ace_root / "eval_main.py"
    if sha256_file(scorer) != manifest.get("officialScorerSha256"):
        raise RuntimeError("ACEBench official scorer drift")
    adapter = Path(str(manifest.get("adapter")))
    if sha256_file(adapter) != manifest.get("adapterSha256"):
        raise RuntimeError("ACEBench scorer adapter drift")
    if manifest.get("historicalResultInput") is not False:
        raise RuntimeError("historical result input is not forbidden")
    if manifest.get("resume") is not False:
        raise RuntimeError("score resume is not forbidden")
    if manifest.get("transformation") != (
        "none; normal FC values passed directly to pinned normal_checker"
    ):
        raise RuntimeError("unexpected native scoring transformation")

    sys.path.insert(0, str(ace_root))
    from model_inference.utils import decode_ast  # noqa: PLC0415

    source_files = manifest.get("sourceFiles")
    if not isinstance(source_files, list) or len(source_files) != 68:
        raise RuntimeError("expected 68 source result files")
    row_totals: dict[str, int] = defaultdict(int)
    seen_categories: dict[str, set[str]] = defaultdict(set)
    category_scores: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    for record in source_files:
        if not isinstance(record, dict):
            raise RuntimeError("invalid source file record")
        path = Path(str(record["path"]))
        if sha256_file(path) != record.get("sha256"):
            raise RuntimeError(f"source result hash drift: {path}")
        rows = load_jsonl(path)
        if len(rows) != record.get("rowCount"):
            raise RuntimeError(f"source result row drift: {path}")
        ids = [row.get("id") for row in rows]
        if any(not isinstance(value, str) for value in ids):
            raise RuntimeError(f"missing task id: {path}")
        if len(ids) != len(set(ids)):
            raise RuntimeError(f"duplicate task id: {path}")
        category = category_from_path(path)
        if category not in ALL_CATEGORIES:
            raise RuntimeError(f"unexpected ACEBench category: {category}")
        language, model = str(record["language"]), str(record["model"])
        if language not in LANGUAGES or model not in ARMS:
            raise RuntimeError(f"unexpected language/model: {language}/{model}")
        if category in NORMAL_CATEGORIES:
            for row in rows:
                decoded = decode_ast(model, row.get("result"))
                if not isinstance(decoded, list) or any(
                    not isinstance(call, dict) for call in decoded
                ):
                    raise RuntimeError(f"invalid native FC value: {path}")

        score_path = (
            scoring_root
            / "score_all"
            / f"score_{language}"
            / model
            / f"data_{category}_score.json"
        )
        score_rows = load_jsonl(score_path)
        if not score_rows:
            raise RuntimeError(f"empty score file: {score_path}")
        metadata = score_rows[0]
        if metadata.get("total_count") != len(rows):
            raise RuntimeError(f"score denominator drift: {score_path}")
        score_key = (
            "end_to_end_accuracy" if category in AGENT_CATEGORIES else "accuracy"
        )
        score = metadata.get(score_key)
        if not isinstance(score, (int, float)) or not 0 <= score <= 1:
            raise RuntimeError(f"invalid score: {score_path}")
        category_scores[language][model][category] = float(score)
        row_totals[model] += len(rows)
        seen_categories[f"{language}/{model}"].add(category)

    if dict(row_totals) != {
        "glm52-native-FC": 2040,
        "glm52-native-plus-FC": 2040,
    }:
        raise RuntimeError(f"ACEBench row coverage drift: {dict(row_totals)}")
    if any(categories != ALL_CATEGORIES for categories in seen_categories.values()):
        raise RuntimeError("ACEBench category coverage drift")
    if len(seen_categories) != 4:
        raise RuntimeError("ACEBench language/arm coverage drift")

    official_by_language: dict[str, dict[str, float]] = {}
    for language in LANGUAGES:
        workbook_path = (
            scoring_root / "score_all" / f"score_{language}" / "result.xlsx"
        )
        workbook_scores = workbook_summaries(workbook_path)
        if set(workbook_scores) != set(ARMS):
            raise RuntimeError(f"workbook arm coverage drift: {workbook_path}")
        for model in ARMS:
            scores = category_scores[language][model]
            normal = round(
                sum(scores[name] for name in NORMAL_CATEGORIES)
                / len(NORMAL_CATEGORIES),
                3,
            )
            special = round(
                sum(scores[name] for name in SPECIAL_CATEGORIES)
                / len(SPECIAL_CATEGORIES),
                3,
            )
            agent = round(
                sum(scores[name] for name in AGENT_CATEGORIES)
                / len(AGENT_CATEGORIES),
                3,
            )
            recomputed = round(
                special * 0.2676 + normal * 0.578 + agent * 0.1545, 3
            )
            if abs(recomputed - workbook_scores[model]) > 1e-12:
                raise RuntimeError(
                    f"official summary drift: {language}/{model} "
                    f"{recomputed} != {workbook_scores[model]}"
                )
        official_by_language[language] = workbook_scores

    macro = {
        model: round(
            sum(official_by_language[language][model] for language in LANGUAGES)
            / len(LANGUAGES),
            4,
        )
        for model in ARMS
    }
    score_files = list((scoring_root / "score_all").rglob("*_score.json"))
    process_files = list((scoring_root / "score_all").rglob("*_process.json"))
    workbook_files = list((scoring_root / "score_all").rglob("result.xlsx"))
    if len(score_files) != 68 or len(process_files) != 8 or len(workbook_files) != 2:
        raise RuntimeError(
            "ACEBench official score tree coverage drift: "
            f"score={len(score_files)} process={len(process_files)} "
            f"workbook={len(workbook_files)}"
        )
    tree_files = score_files + process_files + workbook_files
    summary = {
        "benchmark": "ACEBench native-tool full-population adaptation",
        "categoryFilesPerLanguageArm": 17,
        "officialSummaryByLanguage": official_by_language,
        "officialSummaryMacroAverage": macro,
        "rowsPerArm": dict(row_totals),
        "scoreTreeFileCount": len(tree_files),
        "scoreTreeSha256": canonical_tree_hash(scoring_root, tree_files),
        "scorerCommit": revision,
        "status": "valid",
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, sort_keys=True))


if __name__ == "__main__":
    main()
